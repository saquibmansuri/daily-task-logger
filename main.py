import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

tasks = []

def add_task():
    project = project_entry.get().strip()
    task = task_entry.get().strip()
    task_id = task_id_entry.get().strip()
    hours = hours_entry.get().strip()
    comments = comments_entry.get().strip()

    if not project or not task:
        messagebox.showerror("Error", "Project Name, Task Description and Hours Spent are required!")
        return

    # Validate hours (required, numeric, < 24)
    try:
        hours_val = float(hours)
        if hours_val <= 0 or hours_val >= 24:
            messagebox.showerror("Error", "Hours must be greater than 0 and less than 24!")
            return
    except ValueError:
        messagebox.showerror("Error", "Hours must be a number (e.g., 1 or 1.5)!")
        return

    task_dict = {
        "project": project,
        "task": task,
        "task_id": task_id,
        "hours": hours,
        "comments": comments
    }
    tasks.append(task_dict)

    # Insert into Treeview
    tree.insert("", "end", values=(project, task, task_id, hours, comments))

    # Clear fields
    project_entry.delete(0, tk.END)
    task_entry.delete(0, tk.END)
    task_id_entry.delete(0, tk.END)
    hours_entry.delete(0, tk.END)
    comments_entry.delete(0, tk.END)

def edit_cell(event):
    """Enable editing of a Treeview cell on double-click"""
    selected_item = tree.identify_row(event.y)
    selected_col = tree.identify_column(event.x)

    if not selected_item or selected_col == "#0":
        return

    x, y, width, height = tree.bbox(selected_item, selected_col)
    column_index = int(selected_col.replace("#", "")) - 1

    value = tree.set(selected_item, column=tree["columns"][column_index])

    entry = tk.Entry(tree)
    entry.place(x=x, y=y, width=width, height=height)
    entry.insert(0, value)
    entry.focus()

    def save_edit(event):
        new_value = entry.get().strip()
        col_name = tree["columns"][column_index]

        # Run validations again
        if col_name in ("project", "task") and not new_value:
            messagebox.showerror("Error", f"{col_name.capitalize()} is required!")
            entry.destroy()
            return
        if col_name == "hours":
            try:
                hours_val = float(new_value)
                if hours_val <= 0 or hours_val >= 24:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Hours must be a number between 0 and 24!")
                entry.destroy()
                return

        entry.destroy()
        tree.set(selected_item, column=col_name, value=new_value)

        # Sync back to tasks list
        row_index = tree.index(selected_item)
        tasks[row_index][col_name.lower()] = new_value

    entry.bind("<Return>", save_edit)
    entry.bind("<FocusOut>", save_edit)

def save_to_excel():
    file_name = "tasks.xlsx"
    today = datetime.now().strftime("%d-%m-%Y")

    # Create workbook if not exists
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        headers = ["DATE", "PROJECT", "TASK", "TASK ID", "HOURS", "COMMENTS"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.auto_filter.ref = ws.dimensions
        try:
            wb.save(file_name)
        except PermissionError:
            return False, f"Cannot create {file_name}. Please close it if open and try again."

    # Load workbook
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        for t in tasks:
            ws.append([today, t["project"], t["task"], t["task_id"], t["hours"], t["comments"]])
        ws.auto_filter.ref = ws.dimensions
        wb.save(file_name)
    except PermissionError:
        return False, f"Cannot save {file_name}. Please close it if open and try again."

    return True, "Saved successfully."

def save_tasks():
    if not tasks:
        messagebox.showerror("Error", "No tasks to save!")
        return

    # Save TXT
    today_txt = datetime.now().strftime("%d %B %Y").upper()
    with open("tasks.txt", "a", encoding="utf-8") as f:
        f.write(f"\n*** DATE : {today_txt} ***\n\n")
        for t in tasks:
            f.write(f"Project: {t['project']}\n")
            f.write(f"Task: {t['task']}\n")
            f.write(f"Task ID: {t['task_id']}\n")
            f.write(f"Hours: {t['hours']}\n")
            f.write(f"Comments: {t['comments']}\n\n")

    # Save Excel
    success, msg = save_to_excel()
    if not success:
        messagebox.showerror("Error", msg)
        return

    messagebox.showinfo("Saved", "Tasks saved successfully to TXT and Excel!")
    tasks.clear()
    root.destroy()

# ------------------ UI ------------------
root = tk.Tk()
root.title("Daily Task Logger")

window_width = 750
window_height = 600
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
position_top = int((screen_height/2) - (window_height/2))
position_right = int((screen_width/2) - (window_width/2))
root.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")

bold_font = ("Arial", 10, "bold")

# Input fields
tk.Label(root, text="Project Name (required):", font=bold_font).pack(pady=5)
project_entry = tk.Entry(root, width=50)
project_entry.pack()

tk.Label(root, text="Task Description (required):", font=bold_font).pack(pady=5)
task_entry = tk.Entry(root, width=50)
task_entry.pack()

tk.Label(root, text="Task ID (optional):", font=bold_font).pack(pady=5)
task_id_entry = tk.Entry(root, width=50)
task_id_entry.pack()

tk.Label(root, text="Hours Spent (required):", font=bold_font).pack(pady=5)
hours_entry = tk.Entry(root, width=50)
hours_entry.pack()

tk.Label(root, text="Comments (optional):", font=bold_font).pack(pady=5)
comments_entry = tk.Entry(root, width=50)
comments_entry.pack()

tk.Button(root, text="+ Add Task", command=add_task).pack(pady=10)
tk.Button(root, text="Save & Exit", command=save_tasks).pack(pady=5)

# Task preview (Treeview + scrollbars)
tk.Label(root, text="Tasks Added:", font=bold_font).pack(pady=5)
columns = ("project", "task", "task_id", "hours", "comments")

frame = tk.Frame(root)
frame.pack(fill="both", expand=True, padx=10, pady=10)

tree = ttk.Treeview(frame, columns=columns, show="headings", height=10)
for col in columns:
    tree.heading(col, text=col.upper())
    tree.column(col, width=120, anchor="center")

# Scrollbars
vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
tree.configure(yscroll=vsb.set, xscroll=hsb.set)

tree.grid(row=0, column=0, sticky="nsew")
vsb.grid(row=0, column=1, sticky="ns")
hsb.grid(row=1, column=0, sticky="ew")

frame.grid_rowconfigure(0, weight=1)
frame.grid_columnconfigure(0, weight=1)

tree.bind("<Double-1>", edit_cell)

root.mainloop()
